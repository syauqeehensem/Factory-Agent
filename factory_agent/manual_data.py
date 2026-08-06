"""Technician manual indexing for data/ retrieval.

Builds a lightweight local index from files under the configured technician-doc
folder and supports keyword-based snippet retrieval for the Technician agent.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from .config import settings


@dataclass(frozen=True)
class ManualChunk:
    source: str
    text: str


class ManualIndex:
    """Small in-memory chunk index with keyword-overlap ranking."""

    def __init__(
        self,
        root_dir: str | Path,
        max_pdf_pages: int = 0,
        max_chunks_per_file: int = 0,
        lazy_status: bool = True,
    ) -> None:
        self.root_dir = Path(root_dir)
        self.max_pdf_pages = max(0, int(max_pdf_pages))
        self.max_chunks_per_file = max(0, int(max_chunks_per_file))
        self.lazy_status = bool(lazy_status)
        self.chunks: list[ManualChunk] = []
        self.load_error: str | None = None
        self._loaded = False
        self.search_cache_size = max(0, int(settings.manual_search_cache_size))
        self._search_cache: dict[tuple[str, int], str] = {}
        self._cache_hits = 0
        self._cache_misses = 0

    def ensure_loaded(self) -> None:
        if not self._loaded:
            self.reload()

    def reload(self) -> None:
        self._loaded = True
        self.chunks = []
        self.load_error = None
        self._search_cache.clear()
        self._cache_hits = 0
        self._cache_misses = 0

        if not self.root_dir.exists():
            self.load_error = f"folder not found at {self.root_dir}"
            return

        # Only index technician docs; the data CSVs (status/mtp/yield) are not manuals.
        files = sorted(
            p
            for p in self.root_dir.rglob("*")
            if p.is_file() and p.suffix.lower() in {".pdf", ".xlsx", ".txt", ".md"}
        )
        if not files:
            self.load_error = f"no supported files found under {self.root_dir}"
            return

        for path in files:
            text = self._read_file(path)
            if not text:
                continue
            file_chunks = self._chunk_text(text)
            if self.max_chunks_per_file:
                file_chunks = file_chunks[: self.max_chunks_per_file]
            for idx, chunk in enumerate(file_chunks, start=1):
                source = f"{path.name}#chunk{idx}"
                self.chunks.append(ManualChunk(source=source, text=chunk))

        if not self.chunks:
            self.load_error = (
                "supported files were found but no readable content was extracted; "
                "install optional parsers pypdf/openpyxl for PDF/XLSX support"
            )

    @property
    def is_ready(self) -> bool:
        return self.load_error is None and bool(self.chunks)

    def status_text(self) -> str:
        if not self._loaded and self.lazy_status:
            return (
                "Technician manual index not loaded yet (lazy mode). "
                "It will load on first manual search."
            )
        self.ensure_loaded()
        if self.load_error:
            return f"Technician manual index unavailable: {self.load_error}"
        files = {c.source.split("#", 1)[0] for c in self.chunks}
        return (
            f"Technician manual index ready: {len(files)} file(s), "
            f"{len(self.chunks)} chunks. "
            f"Search cache: {len(self._search_cache)} item(s), "
            f"hits={self._cache_hits}, misses={self._cache_misses}."
        )

    def search(self, query: str, top_k: int = 3) -> str:
        self.ensure_loaded()
        if self.load_error:
            return f"Technician manual index unavailable: {self.load_error}"
        if not self.chunks:
            return "Technician manual index is empty."

        normalized_query = self._normalize_query(query)
        k = max(1, int(top_k))
        cache_key = (normalized_query, k)
        cached = self._search_cache.get(cache_key)
        if cached is not None:
            self._cache_hits += 1
            return cached
        self._cache_misses += 1

        terms = self._terms(query)
        if not terms:
            return "Please include a clearer manual question (error code, tool id, or symptom)."

        scored: list[tuple[int, ManualChunk]] = []
        for chunk in self.chunks:
            score = self._score_terms(chunk.text, terms)
            if score > 0:
                scored.append((score, chunk))

        if not scored:
            result = (
                "I could not find a close manual match for that question. "
                "Try adding the exact symptom, station, or error keyword."
            )
            self._remember_search(cache_key, result)
            return result

        scored.sort(key=lambda item: item[0], reverse=True)
        rows = ["Top technician manual snippets:"]
        for rank, (score, chunk) in enumerate(scored[:k], start=1):
            rows.append(
                f"{rank}. [{chunk.source}] score={score}\n"
                f"   {self._trim(chunk.text, 260)}"
            )
        result = "\n".join(rows)
        self._remember_search(cache_key, result)
        return result

    @staticmethod
    def _normalize_query(query: str) -> str:
        return " ".join((query or "").strip().lower().split())

    def _remember_search(self, key: tuple[str, int], value: str) -> None:
        if self.search_cache_size <= 0:
            return
        if key in self._search_cache:
            self._search_cache.pop(key, None)
        elif len(self._search_cache) >= self.search_cache_size:
            oldest = next(iter(self._search_cache), None)
            if oldest is not None:
                self._search_cache.pop(oldest, None)
        self._search_cache[key] = value

    @staticmethod
    def _terms(text: str) -> set[str]:
        return {t for t in re.findall(r"[A-Za-z0-9_\-/\.]{2,}", text.lower())}

    @classmethod
    def _score_terms(cls, text: str, terms: set[str]) -> int:
        hay = cls._terms(text)
        return sum(2 if term in hay and any(ch.isdigit() for ch in term) else 1 for term in terms if term in hay)

    @staticmethod
    def _chunk_text(text: str, size: int = 1200, overlap: int = 120) -> list[str]:
        normalized = " ".join(text.split())
        if not normalized:
            return []
        chunks: list[str] = []
        start = 0
        while start < len(normalized):
            end = min(len(normalized), start + size)
            snippet = normalized[start:end].strip()
            if snippet:
                chunks.append(snippet)
            if end >= len(normalized):
                break
            start = max(0, end - overlap)
        return chunks

    @staticmethod
    def _trim(text: str, max_chars: int) -> str:
        if len(text) <= max_chars:
            return text
        clipped = text[: max_chars - 3].rstrip()
        if " " in clipped:
            clipped = clipped.rsplit(" ", 1)[0]
        return f"{clipped}..."

    def _read_file(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".txt", ".md", ".csv"}:
            return self._read_text(path)
        if suffix == ".pdf":
            return self._read_pdf(path)
        if suffix == ".xlsx":
            return self._read_xlsx(path)
        return ""

    @staticmethod
    def _read_text(path: Path) -> str:
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            return ""

    def _read_pdf(self, path: Path) -> str:
        try:
            from pypdf import PdfReader
        except Exception:
            return ""
        try:
            reader = PdfReader(str(path))
            out: list[str] = []
            page_count = len(reader.pages)
            limit = page_count
            if self.max_pdf_pages > 0:
                limit = min(page_count, self.max_pdf_pages)
            for idx in range(limit):
                try:
                    txt = reader.pages[idx].extract_text() or ""
                except Exception:
                    continue
                if txt:
                    out.append(txt)
            return "\n".join(out)
        except Exception:
            return ""

    @staticmethod
    def _read_xlsx(path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except Exception:
            return ""

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            out: list[str] = []
            for ws in wb.worksheets:
                out.append(f"Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                    if vals:
                        out.append(" | ".join(vals))
            wb.close()
            return "\n".join(out)
        except Exception:
            return ""


MANUAL_INDEX = ManualIndex(
    root_dir=settings.technician_docs_dir,
    max_pdf_pages=settings.manual_max_pdf_pages,
    max_chunks_per_file=settings.manual_max_chunks_per_file,
    lazy_status=settings.manual_status_lazy,
)
